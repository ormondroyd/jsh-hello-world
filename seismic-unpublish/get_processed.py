from openpyxl import load_workbook
import sys, json

wb = load_workbook(sys.argv[1])
ws = wb.active

done = []
for row in ws.iter_rows(min_row=2):
    if row[1].value in ('Unpublished', 'Already unpublished', 'Expired'):
        done.append(str(row[0].value or '').strip())

print(json.dumps(done))
