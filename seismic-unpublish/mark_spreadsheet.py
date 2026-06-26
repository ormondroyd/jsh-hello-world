from openpyxl import load_workbook
from datetime import datetime
import sys

wb = load_workbook(sys.argv[1])
ws = wb.active

if ws.cell(1, 2).value != 'Status':
    ws.cell(1, 2).value = 'Status'
    ws.cell(1, 3).value = 'Timestamp'

name = sys.argv[2]
status = sys.argv[3]
for row in ws.iter_rows(min_row=2):
    if str(row[0].value or '').strip() == name:
        row[1].value = status
        row[2].value = datetime.now().strftime('%Y-%m-%d %H:%M')
        break

wb.save(sys.argv[1])
